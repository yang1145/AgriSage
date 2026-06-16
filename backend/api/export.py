import os
import shutil
from datetime import date
from io import BytesIO

from flask import Blueprint, request, jsonify, current_app, send_file
from openpyxl import load_workbook

from extensions import db
from models import Plot, PlantingCycle, User
from utils.export_helpers import export_to_excel, export_to_pdf
from api.auth import token_required, role_required

export_bp = Blueprint('export', __name__)


@export_bp.route('/excel', methods=['POST'])
@token_required
def export_excel(current_user):
    """导出地块汇总到Excel"""
    data = request.get_json() or {}
    plot_ids = data.get('plot_ids', [])

    if plot_ids:
        plots = Plot.query.filter(Plot.id.in_(plot_ids)).all()
    else:
        if current_user.role == 'owner':
            plots = Plot.query.filter_by(user_id=current_user.id).all()
        else:
            plots = Plot.query.filter_by(user_id=current_user.id).all()

    rows = []
    for p in plots:
        rows.append(p.to_dict())

    headers = [
        ('id', 'ID'),
        ('name', '地块名称'),
        ('area', '面积(亩)'),
        ('elevation', '海拔(m)'),
        ('slope', '坡度(度)'),
        ('slope_aspect', '坡向'),
        ('soil_type', '土壤类型'),
        ('soil_ph', '土壤pH'),
        ('organic_matter', '有机质'),
        ('soil_depth', '土层深度(cm)'),
        ('status', '状态'),
        ('township', '乡镇'),
        ('created_at', '创建时间'),
    ]

    buffer = export_to_excel(rows, headers, filename='地块汇总')

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='地块汇总.xlsx',
    )


@export_bp.route('/pdf/<int:plot_id>', methods=['POST'])
@token_required
def export_pdf(current_user, plot_id):
    """导出单个地块PDF档案"""
    plot = Plot.query.get(plot_id)
    if not plot:
        return jsonify({'message': '地块不存在'}), 404

    # 基本信息
    plot_data = plot.to_dict()

    # 农事时间线
    cycles = PlantingCycle.query.filter_by(plot_id=plot_id).all()
    timeline_rows = []
    for cycle in cycles:
        timeline_rows.append({
            '日期': cycle.plant_date.isoformat() if cycle.plant_date else '',
            '类型': f'种植({cycle.cycle_type})',
            '详情': f'品种ID:{cycle.variety_id or "无"}',
        })
        for r in cycle.farming_records_fertilization.all():
            timeline_rows.append({
                '日期': r.date.isoformat() if r.date else '',
                '类型': '施肥',
                '详情': f'{r.fertilizer_name or ""} {r.amount or ""}kg/亩',
            })
        for r in cycle.farming_records_irrigation.all():
            timeline_rows.append({
                '日期': r.date.isoformat() if r.date else '',
                '类型': '灌溉',
                '详情': f'{r.method or ""} {r.water_amount or ""}m³/亩',
            })
        for r in cycle.farming_records_pest.all():
            timeline_rows.append({
                '日期': r.discovery_date.isoformat() if r.discovery_date else '',
                '类型': '病虫害',
                '详情': f'{r.pest_type or ""} 严重程度:{r.severity or ""}',
            })
        for r in cycle.farming_records_harvest.all():
            timeline_rows.append({
                '日期': (r.actual_date or r.planned_date).isoformat() if (r.actual_date or r.planned_date) else '',
                '类型': '收获',
                '详情': f'产量:{r.yield_tons or ""}吨',
            })

    headers = [
        ('日期', '日期'),
        ('类型', '类型'),
        ('详情', '详情'),
    ]

    title = f'地块档案 - {plot.name}'
    buffer = export_to_pdf(timeline_rows, headers, filename=f'plot_{plot_id}', title=title)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'{plot.name}_档案.pdf',
    )


@export_bp.route('/import-excel', methods=['POST'])
@role_required('coop_admin')
def import_excel(current_user):
    """导入Excel数据（仅合作社管理员）"""
    if 'file' not in request.files:
        return jsonify({'message': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': '未选择文件'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'message': '仅支持Excel文件(.xlsx/.xls)'}), 400

    try:
        wb = load_workbook(file)
        ws = wb.active

        imported_count = 0
        errors = []

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = dict(zip(headers, row))

                # 根据表头判断导入类型
                if '地块名称' in row_data and row_data['地块名称']:
                    plot = Plot(
                        user_id=current_user.id,
                        name=str(row_data.get('地块名称', '')),
                        area=row_data.get('面积(亩)') or row_data.get('面积'),
                        elevation=row_data.get('海拔(m)') or row_data.get('海拔'),
                        soil_type=row_data.get('土壤类型'),
                        township=row_data.get('乡镇'),
                        status=row_data.get('状态', '闲置'),
                    )
                    db.session.add(plot)
                    imported_count += 1

            except Exception as e:
                errors.append(f'第{row_idx}行: {str(e)}')
                continue

        db.session.commit()

        result = {'message': f'成功导入 {imported_count} 条记录', 'imported': imported_count}
        if errors:
            result['errors'] = errors

        return jsonify(result), 200

    except Exception as e:
        return jsonify({'message': f'导入失败: {str(e)}'}), 500


@export_bp.route('/backup/download', methods=['GET'])
@role_required('owner')
def download_backup(current_user):
    """下载数据库备份文件（仅户主）"""
    db_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
    if db_uri.startswith('sqlite:///'):
        db_path = db_uri.replace('sqlite:///', '')
        if not os.path.isabs(db_path):
            db_path = os.path.join(current_app.root_path, '..', db_path)

        if not os.path.exists(db_path):
            return jsonify({'message': '数据库文件不存在'}), 404

        return send_file(
            db_path,
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name='agrisage_backup.db',
        )

    return jsonify({'message': '仅支持SQLite数据库备份'}), 400


@export_bp.route('/template', methods=['GET'])
@token_required
def download_template(current_user):
    """下载Excel导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '地块导入模板'

    headers = ['地块名称', '面积(亩)', '海拔(m)', '坡度(度)', '坡向', '土壤类型', '土壤pH', '有机质', '土层深度(cm)', '状态', '乡镇']
    ws.append(headers)

    # 样式
    header_fill = PatternFill(start_color='1a2332', end_color='1a2332', fill_type='solid')
    header_font = Font(color='00f0ff', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 示例行
    ws.append(['示例地块A', '5.2', '120', '3', '东南', '红壤', '5.5', '2.8', '80', '闲置', '某某乡'])

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='AgriSage_导入模板.xlsx',
    )


@export_bp.route('/backup/restore', methods=['POST'])
@role_required('owner')
def restore_backup(current_user):
    """从备份文件恢复数据库（仅户主）"""
    if 'file' not in request.files:
        return jsonify({'message': '未上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': '未选择文件'}), 400

    if not file.filename.endswith('.db'):
        return jsonify({'message': '仅支持.db文件'}), 400

    db_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
    if not db_uri.startswith('sqlite:///'):
        return jsonify({'message': '仅支持SQLite数据库恢复'}), 400

    db_path = db_uri.replace('sqlite:///', '')
    if not os.path.isabs(db_path):
        db_path = os.path.join(current_app.root_path, '..', db_path)

    try:
        # 备份当前数据库
        backup_path = db_path + '.bak'
        if os.path.exists(db_path):
            shutil.copy2(db_path, backup_path)

        # 保存上传的文件
        file.save(db_path)

        return jsonify({'message': '数据库恢复成功，请重启应用'}), 200

    except Exception as e:
        # 恢复失败时还原备份
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, db_path)
        return jsonify({'message': f'恢复失败: {str(e)}'}), 500
