import io
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors


def export_to_excel(data, headers, filename='export'):
    """
    Export data to an Excel file.

    Args:
        data: List of dicts, each dict represents a row.
        headers: List of (key, label) tuples defining columns.
        filename: Output filename (without extension).

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = filename[:31]  # Excel sheet name max 31 chars

    # Write header row
    for col_idx, (_, label) in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=label)

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, (key, _) in enumerate(headers, 1):
            value = row_data.get(key, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_to_pdf(data, headers, filename='export', title='数据报表'):
    """
    Export data to a PDF file.

    Args:
        data: List of dicts, each dict represents a row.
        headers: List of (key, label) tuples defining columns.
        filename: Output filename (without extension).
        title: Report title.

    Returns:
        BytesIO buffer containing the PDF file.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles['Title']
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph('', styles['Normal']))  # Spacer

    # Build table data
    table_data = [[label for _, label in headers]]
    for row_data in data:
        row = [str(row_data.get(key, '')) for key, _ in headers]
        table_data.append(row)

    # Calculate available width
    available_width = A4[0] - 30 * mm
    col_count = len(headers)
    col_width = available_width / col_count if col_count > 0 else available_width

    table = Table(table_data, colWidths=[col_width] * col_count)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
